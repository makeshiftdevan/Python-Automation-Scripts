import os, time, re
import openpyxl
from openpyxl.styles import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import formatting, styles
import pandas as pd
import xlsxwriter
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import datetime
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 

#Changes current working directory (CWD) to this path, else FileNotFoundError
os.chdir(r'C:\Users\...\downloads') 

#location of chromedriver
chromedriver = 'C:/Users/.../chromedriver.exe'

#tells selenium where to find the driver to open Chrome
driver = webdriver.Chrome(chromedriver) 

report = Workbook(r'C:\Users\...\downloads\EdgenuityReport.xlsx')


#To log in
def log_in():
    #to minimize window
    driver.minimize_window()
    driver.get('https://auth.edgenuity.com/Login/Login/Educator')
    driver.find_element_by_id('LoginUsername').send_keys('d***t')
    driver.find_element_by_id ('LoginPassword').send_keys('*****')
    driver.find_element_by_id('LoginSubmit').click()
    time.sleep(3)
    try:
        driver.find_element_by_id("lbContinue").click()
    except:
        pass
    time.sleep(5)
    
    #clicks 'dashboard'
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_pnlMain"]/table[1]/tbody/tr/td[2]/a').click()  
    time.sleep(5)

    #course filters button opens a popup
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(3)
log_in()

#Physics
def report0():
    time.sleep(5)
    
     #clicks the current option in the drop down of the popup
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click()
    time.sleep(5)

     #selects the class option
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[91]').click() 
    time.sleep(5)

    #apply
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(5)

    #exports as csv
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 1 of 6 downloaded')
    time.sleep(5)
report0()

#Physical Science
def report1():
    dataframe = []
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[203]').click() 
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 2 of 6 downloaded')
    time.sleep(3)
report1()

#AP Physics Mechanics Review
def report2():
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[202]').click()
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 3 of 6 downloaded')
    time.sleep(5)
report2()

#AP Physics E&M
def report3():
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[159]').click()  
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 4 of 6 downloaded')
    time.sleep(5)
report3()

#IB Year 1 Review 1
def report4():
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click() 
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[200]').click() 
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(10)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 5 of 6 downloaded')
    time.sleep(5)
report4()

#IB Year 1 Reviewe 2
def report5():
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnCourseFilters').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]').click()  
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_ddlCourses"]/option[201]').click() 
    time.sleep(5)
    driver.find_element_by_id('ctl00_conBody_btnApplyCourse').click() 
    time.sleep(10)
    driver.find_element_by_xpath('//*[@id="ctl00_conBody_btnExport"]').click()
    print('Report 6 of 6 downloaded')
    time.sleep(5)
report5()

print("Reading reports...")



#compile all reports into 1 csv



report = Workbook(r'C:\Users\...\downloads\EdgenuityReport.xlsx')

report.save(r'C:\Users\...\downloads\EdgenuityReport.xlsx')

wb = load_workbook('EdgenuityReport.xlsx')

myDir = r'C:\Users\...\downloads'

# a dumping groups for data
dataframe = [] 

csv_files = os.listdir( r'C:\Users\...\downloads')

red_color_font = '9c0103'
red_font = styles.Font(size=14, bold=True, color=red_color_font)

#first fill in dataframe

#read each csv file
df0 = pd.read_csv(r'C:\Users\...\downloads\Dashboard.csv')

#send info, df0, from csv to dataframe
dataframe.append(df0)
df1= pd.read_csv(r'C:\Users\...\downloads\Dashboard (1).csv')
dataframe.append(df1)
df2 = pd.read_csv(r'C:\Users\...\downloads\Dashboard (2).csv')
dataframe.append(df2)
df3 = pd.read_csv(r'C:\Users\...\downloads\Dashboard (3).csv')
dataframe.append(df3)
df4 = pd.read_csv(r'C:\Users\...\downloads\Dashboard (4).csv')
dataframe.append(df4)
df5 = pd.read_csv(r'C:\Users\...\downloads\Dashboard (5).csv')
dataframe.append(df5)


#Now, write in dataframe to excel
writer = pd.ExcelWriter(r'C:\Users\devan\downloads\EdgenuityReport.xlsx', engine='xlsxwriter')

#must have this because you cannot format a 'Workbook' (ie report).
workbook = writer.book 
format1 = workbook.add_format({'bold': True, 'font_color': 'red'})


df0.to_excel(writer, sheet_name="Physics", columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))

#allows for conditional formating of the selected worksheet
worksheet = writer.sheets['Physics'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


df1.to_excel(writer, sheet_name="PhysicalScience", columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))
worksheet = writer.sheets['PhysicalScience'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


df2.to_excel(writer, sheet_name='APC_Mech', columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))
worksheet = writer.sheets['APC_Mech'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


df3.to_excel(writer, sheet_name='APC_E&M', columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))
worksheet = writer.sheets['APC_E&M'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


df4.to_excel(writer, sheet_name='IB.1', columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))
worksheet = writer.sheets['IB.1'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


df5.to_excel(writer, sheet_name='IB.2', columns = ("External User ID", "Last Name", "First Name", "Grade", "Days Since Last Action", "Target Completion", "Progress"))
worksheet = writer.sheets['IB.2'] 
worksheet.conditional_format('E2:E50', {'type':'cell', 'criteria': '>=', 'value':    3, 'format': format1})


writer.save()

#delete all files once compiled into 1 report
os.remove('Dashboard.csv')
os.remove('Dashboard (1).csv')
os.remove('Dashboard (2).csv')
os.remove('Dashboard (3).csv')
os.remove('Dashboard (4).csv')
os.remove('Dashboard (5).csv')

print('Report Compiled, preparing email.')

#emailing report

fromaddr = "enterFromEmailAddressHere"
toaddr = "enterToEmailAddressHere"
   
# instance of MIMEMultipart 
msg = MIMEMultipart() 
  
# storing the senders email address   
msg['From'] = fromaddr 
  
# storing the receivers email address  
msg['To'] = toaddr 
  
# storing the subject
time_now = datetime.datetime.now().isoformat()
msg['Subject'] = "Edgenuity Report" + time_now[0:10]
  
# string to store the body of the mail 
body = r"Attached you will find today's report"
  
# attach the body with the msg instance 
msg.attach(MIMEText(body, 'plain')) 
  
# open the file to be sent  
filename = "EdgenuityReport.xlsx"
attachment = open(r'C:\Users\devan\downloads\EdgenuityReport.xlsx', "rb") 
  
# instance of MIMEBase and named as p 
p = MIMEBase('application', 'octet-stream') 
  
# To change the payload into encoded form 
p.set_payload((attachment).read()) 
  
# encode into base64 
encoders.encode_base64(p) 
   
p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
  
# attach the instance 'p' to instance 'msg' 
msg.attach(p) 
  
# creates SMTP session 
s = smtplib.SMTP('smtp.gmail.com', 587) 
  
# start TLS for security 
s.starttls() 
  
# Authentication 
s.login(fromaddr, "fromAddressPasswordHere") 
  
# Converts the Multipart msg into a string 
text = msg.as_string() 
  
# sending the mail 
s.sendmail(fromaddr, toaddr, text) 
  
# terminating the session 
s.quit()

print('Edgenuity Report Successfully Generated!')

