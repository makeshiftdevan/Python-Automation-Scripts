import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import datetime
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 

wb = load_workbook(filename = r'C:\Users\...\Downloads\EdgenuityReport.xlsx')

grades = []

doneness = []

#iterates through each sheet in a workbook and reads specific column(s) and adds to list above
for sheet in wb:
    def get_grades():
        for row in sheet.iter_cols(min_col = 5, max_col = 5, min_row= 2, max_row =30, values_only = True):
        
            # using List Comprehension to remove None values in list by checking for True values
            each_student = [i for i in row if i]
            str(each_student)
            grades.append(each_student)
    get_grades()
    

    def percent_complete():
        #percent completion for each student
        for rows in sheet.iter_cols(min_col = 8, max_col = 8, min_row = 2, max_row = 35, values_only = True):
            percent = [i for i in rows if i]
            str(percent)
            doneness.append(percent)
    percent_complete()
    


#lists cannot use .replace.  BUT if you iterate them (for _ in _), they are treated as strings. if the list contains a list, iterate twice.
grades =[[x.replace('%',' ') for x in l] for l in grades]
  #input_grades  [[input_box.replace('*', [for i in grades])]]


doneness[1].insert(1, 0)



#Fire up Chrome
chromedriver = 'C:/Users/.../chromedriver.exe'
driver = webdriver.Chrome(chromedriver)
driver.maximize_window()

#Log into Focus
def log_in():
    driver.get('https://focus.oneclay.net/focus/')
    driver.find_element_by_xpath('//*[@id="username-input"]').send_keys('dsskapetis')
    driver.find_element_by_xpath('//*[@id="main-login-form"]/input[2]').send_keys('Wendell3!')
    driver.find_element_by_xpath('//*[@id="main-login-form"]/div[1]/div[2]/button').click()
    time.sleep(3)
    if driver.find_element_by_xpath('//*[@id="computer-name-form"]/input'):
        driver.find_element_by_xpath('//*[@id="computer-name-form"]/input').send_keys('Egenuity Grading Bot')
        driver.find_element_by_xpath('//*[@id="computer-name-form"]/div/button').click()
        time.sleep(3)
    else:
        time.sleep(3)
log_in()
        

def prepare_Focus():
    time.sleep(3)
    #Grades tab
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/nav/div[2]/button').click()
    time.sleep(3)
    #Gradebook from dropdown
    driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/nav/div[2]/div/div/div[1]/div/a[1]').click()
    time.sleep(5)
    #turning portion of list into an integer to be used as input later
    new_grade0 = ((grades[3])[1])
    new_grade1 = int(float((new_grade0)))
    time.sleep(2)
    #input text finder
    input_box = driver.find_element_by_css_selector('#DTI\[301908\]\[6212517\]')
    #implicit wait for input element to become visible
    driver.implicitly_wait(10)
    #linking actions together, sleep inserted to allow for site to process actions
    ActionChains(driver).move_to_element(input_box).click(input_box).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()
    time.sleep(1)
    ActionChains(driver).send_keys(new_grade1).send_keys(Keys.LEFT).send_keys(Keys.LEFT).send_keys(Keys.LEFT).perform()
    time.sleep(3)
prepare_Focus()

i = 0

def enter_APMech_grades():
    #class list dropdown in top right
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/form/ul/li[4]/select').click()
    #Selection of AP Physics Mech
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/form/ul/li[4]/select/optgroup[1]/option[2]').click()
##    time.sleep(3)
##    #css_selector is much more reliable as a finder than xpath
##    first_grade = driver.find_element_by_css_selector('#DTI\[240546\]\[6212519\]')
##    ActionChains(driver).move_to_element(first_grade).click(first_grade).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()
##
##    #iterates over the selected sheet's grade list
##    for i in range(len(grades[2])):
##        #pulls out the indexed grade from the list
##        time.sleep(3)
##        apMech= ((grades[2][i]))
##        #changes the list to an int to be input
##        apMech1 = int(float((apMech)))
##        #inputs the grade and presses down to go to next student
##        ActionChains(driver).send_keys(apMech1).send_keys(Keys.DOWN).perform()
##        ++i
##        time.sleep(3)
##
##    #moves to next assignment input box
##    time.sleep(3)
##    ActionChains(driver).send_keys(Keys.RIGHT).send_keys(Keys.RIGHT).perform()
##
##    #iterates over the selected sheet's doneness list
##    for i in range(len(doneness[2])):
##        time.sleep(2)
##        #pulls out the indexed percent from the list
##        complete= ((doneness[2][i]))
##        #changes the list to an int to be input
##        complete1 = int(float((complete)))
##        #inputs the progress amount and presses down to go to next student
##        ActionChains(driver).send_keys(complete1).send_keys(Keys.DOWN).perform()
##        ++i
##        time.sleep(3)
##        
##enter_APMech_grades()

def enter_Physical_Science():
    #class list dropdown in top right
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/form/ul/li[4]/select').click()
    #Selection of Physical Science
    driver.find_element_by_xpath('//*[@id="Top_Right"]/li[4]/select/optgroup[1]/option[9]').click()
    time.sleep(3)
    first_grade = driver.find_element_by_css_selector('#DTI\[285195\]\[6206647\] > div.jsDTISpan.jsDTIVisible > span')
    ActionChains(driver).move_to_element(first_grade).click(first_grade).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()

    #iterates over the selected sheet's grades list
    for i in range(len(grades[1])):
        #pulls out the indexed percent from the list
        ps= ((grades[1][i]))
        #changes the list to an int to be input
        ps1 = int(float((ps)))
        #inputs the progress amount and presses down to go to next student
        ActionChains(driver).send_keys(ps1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(5)

    #moves to next assignment input box
    time.sleep(3)
    ActionChains(driver).send_keys(Keys.RIGHT).send_keys(Keys.RIGHT).perform()

    #iterates over the selected sheet's grade list
    for i in range(len(doneness[1])):
        #pulls out the indexed grade from the list
        complete= ((doneness[1][i]))
        #changes the list to an int to be input
        complete1 = int(float(complete))
        time.sleep(2)
        #inputs the grade and presses down to go to next student
        ActionChains(driver).send_keys(complete1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(3)

        
enter_Physical_Science()


def enter_IB1():
    #class list dropdown in top right
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/form/ul/li[4]/select').click()
    #Selection of IB
    driver.find_element_by_xpath('//*[@id="Top_Right"]/li[4]/select/optgroup[1]/option[12]').click()
    time.sleep(3)
    #css_selector is much more reliable as a finder than xpath
    first_grade = driver.find_element_by_css_selector('#DTI\[238920\]\[6215232\] > div.jsDTISpan.jsDTIVisible')
    driver.implicitly_wait(10)
    ActionChains(driver).move_to_element(first_grade).click(first_grade).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()

    #iterates over the selected sheet's doneness list
    for i in range(len(grades[4])):
        #pulls out the indexed percent from the list
        ps= ((grades[4][i]))
        #changes the list to an int to be input
        ps1 = int(float((ps)))
        time.sleep(2)
        #inputs the progress amount and presses down to go to next student
        ActionChains(driver).send_keys(ps1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(3)

    #moves to next assignment input box
    time.sleep(3)
    ActionChains(driver).send_keys(Keys.RIGHT).send_keys(Keys.RIGHT).perform()

    #iterates over the selected sheet's grade list
    for i in range(len(doneness[4])):
        #pulls out the indexed grade from the list
        complete= ((doneness[4][i]))
        #changes the list to an int to be input
        complete1 = int(float(complete))
        #inputs the grade and presses down to go to next student
        ActionChains(driver).send_keys(complete1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(3)

        
enter_IB1()


def enter_physics():
    #class list dropdown in top right
    driver.find_element_by_xpath('/html/body/div[1]/header/div[3]/form/ul/li[4]/select').click()
    #Selection of physics
    driver.find_element_by_xpath('//*[@id="Top_Right"]/li[4]/select/optgroup[1]/option[14]').click()
    time.sleep(3)
    #css_selector is much more reliable as a finder than xpath
    first_grade = driver.find_element_by_css_selector('#DTI\[240318\]\[6206663\] > div.jsDTISpan.jsDTIVisible > span')
    ActionChains(driver).move_to_element(first_grade).click(first_grade).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).send_keys(Keys.BACKSPACE).perform()

    #iterates over the selected sheet's doneness list
    for i in range(len(grades[0])):
        #pulls out the indexed percent from the list
        ps= ((grades[0][i]))
        time.sleep(2)
        #changes the list to an int to be input
        ps1 = int(float((ps)))
        #inputs the progress amount and presses down to go to next student
        ActionChains(driver).send_keys(ps1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(3)

    #moves to next assignment input box
    time.sleep(3)
    ActionChains(driver).send_keys(Keys.RIGHT).send_keys(Keys.RIGHT).perform()

    #iterates over the selected sheet's grade list
    for i in range(len(doneness[0])):
        #pulls out the indexed grade from the list
        complete= ((doneness[0][i]))
        #changes the list to an int to be input
        complete1 = int(float(complete))
        #inputs the grade and presses down to go to next student
        ActionChains(driver).send_keys(complete1).send_keys(Keys.DOWN).perform()
        ++i
        time.sleep(3)

        
enter_physics()

def email_confirmation ():
    fromaddr = "enterFromEmailAddresHere"
    toaddr = "enterToEmailAddresHere"
   
# instance of MIMEMultipart 
    msg = MIMEMultipart() 
  
# storing the senders email address   
    msg['From'] = fromaddr 
  
# storing the receivers email address  
    msg['To'] = toaddr 
  
# storing the subject
    time_now = datetime.datetime.now().isoformat()
    msg['Subject'] = "Grades Entered" + time_now[0:10]
  
# string to store the body of the mail 
body = r"All grades have been updated using the most recent Edgenuity Report."
  
# attach the body with the msg instance 
msg.attach(MIMEText(body, 'plain')) 
  
# instance of MIMEBase and named as p 
p = MIMEBase('application', 'octet-stream') 
  
# encode into base64 
#encoders.encode_base64(p) 
   
p.add_header('Content-Disposition', "attachment;None") 
  
# creates SMTP session 
s = smtplib.SMTP('smtp.gmail.com', 587) 
  
# start TLS for security 
s.starttls() 
  
# Authentication 
s.login(fromaddr, "enterFromEmailAddressPasswordHere") 
  
# Converts the Multipart msg into a string 
text = msg.as_string() 
  
# sending the mail 
s.sendmail(fromaddr, toaddr, text) 
  
# terminating the session 
s.quit()


email_confirmation()

